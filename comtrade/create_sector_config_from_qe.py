"""
create_sector_config_from_qe.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Generates a pre-filled sector_config.xlsx from the QE/PC HS mapping file.
Also supports fully manual product group creation.

USAGE
─────
  List all available QE groups:
      python create_sector_config_from_qe.py --list

  Auto-generate config from a single QE group:
      python create_sector_config_from_qe.py --group "ELECTRONIC GOODS"

  Combine MULTIPLE QE groups into one sector config (use semicolons to separate group names):
      python create_sector_config_from_qe.py --groups "RMG OF ALL TEXTILES;CARPET;JUTE MFG. INCLUDING FLOOR COVERING" --sector "Textiles & Apparel"

  Filter to specific HS chapters only:
      python create_sector_config_from_qe.py --groups "..." --sector "..." --hs_filter "50,51,52,53,54,55,56,57,58,59,60,61,62,63"

  Search for a group by keyword:
      python create_sector_config_from_qe.py --keyword "leather"

  Generate a blank config for manual entry:
      python create_sector_config_from_qe.py --manual --sector "Toys & Games"

  Combine QE group + manual custom codes:
      python create_sector_config_from_qe.py --group "ELECTRONIC GOODS" --add_codes "9031,9032,9033"

OPTIONS
───────
  --group        Single exact QE group name (use --list to see all options)
  --groups       Comma-separated list of QE groups to combine into one sector
  --sector       Sector name (required with --groups or --manual)
  --keyword      Search keyword — finds all matching QE groups
  --manual       Generate a blank config for fully manual entry
  --flow         Which Comtrade flow sheet to use: export (default) or import
  --hs_level     HS code granularity: 2 (chapter), 6 (sub-heading), 8 (ITCHS). Default: 2
  --hs_filter    Comma-separated HS chapters to include (e.g. "50,51,52"). Excludes all others.
  --add_codes    Comma-separated extra HS codes to add alongside QE auto-fill
  --out          Output file name (default: auto-generated from sector name)
  --mapping      Path to QE/PC mapping file (default: auto-detected)

MAPPING FILE
────────────
  Default: QE_PC_HS_Mapping_2025-26_Final_Sent.xlsx  (same folder as this script)
  Sheet EXPORT is used for --flow export (default)
  Sheet IMPORT is used for --flow import

WHAT IT CREATES
───────────────
  A filled sector_config.xlsx with these sheets:
  • sector_details  — pre-filled with sector name, slug, date
  • hs_codes        — HS codes from QE/PC mapping (or blank for manual)
  • sector_buckets  — PC Groups become buckets automatically
  • years           — 2014–2023 default
  • flows           — 4 standard trade flows
  • settings        — pipeline settings
  • custom_groups   — ADD YOUR OWN HS codes here (merged on next run)

  After editing custom_groups, re-run with --merge to fold custom codes into hs_codes.

EXAMPLES
────────
  python create_sector_config_from_qe.py --list
  python create_sector_config_from_qe.py --group "ELECTRONIC GOODS"
  python create_sector_config_from_qe.py --keyword "leather"
  python create_sector_config_from_qe.py --group "GEMS AND JEWELLERY" --hs_level 6
  python create_sector_config_from_qe.py --manual --sector "Toys and Games"
  python create_sector_config_from_qe.py --group "DRUGS AND PHARMACEUTICALS" --add_codes "3001,3002"
"""

import argparse
import os
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# COLOUR PALETTE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DARK   = "1B3A5C"
MID    = "2874A6"
LIGHT  = "D6EAF8"
GREEN  = "E8F8E8"
ORANGE = "FFF3E0"
WHITE  = "FFFFFF"
GOLD   = "FFF9C4"
GRAY   = "F2F2F2"

SCRIPT_DIR = Path(__file__).parent.resolve()
DEFAULT_MAPPING = SCRIPT_DIR / "QE_PC_HS_Mapping_2025-26_Final_Sent.xlsx"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STYLE HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def thin_border():
    t = Side(border_style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def style_cell(ws, r, c, value, bold=False, fill=None, italic=False,
               font_size=10, wrap=False, align="left", color="000000"):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = Font(bold=bold, size=font_size, name="Calibri",
                     italic=italic, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin_border()
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell

def write_title(ws, row, text, ncols, fill=DARK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, name="Calibri", color=WHITE)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24

def write_header(ws, row, cols, fill=DARK):
    """cols = list of (label, width)"""
    for i, (label, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(bold=True, size=10, name="Calibri", color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 28

def note_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(italic=True, size=9, name="Calibri", color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAPPING FILE READER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def load_mapping(mapping_path, flow="export"):
    """
    Returns a dict:
      {
        "QE Group Name": {
          "PC Group Name": [
            {"itchs": "57011010", "unit": "SQM", "description": "HAND-MADE"},
            ...
          ]
        }
      }
    """
    sheet_name = "EXPORT" if flow.lower() == "export" else "IMPORT"
    wb = openpyxl.load_workbook(mapping_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  ✗  Sheet '{sheet_name}' not found in {mapping_path.name}")
        print(f"     Available sheets: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet_name]

    mapping = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        qe_group  = str(row[0]).strip() if row[0] else None
        pc_group  = str(row[1]).strip() if row[1] else None
        itchs     = str(row[3]).strip() if row[3] else None
        unit      = str(row[4]).strip() if row[4] else ""
        desc      = str(row[5]).strip() if row[5] else ""

        if not qe_group or not itchs or itchs == "None":
            continue

        if qe_group not in mapping:
            mapping[qe_group] = {}
        pc = pc_group or "Uncategorised"
        if pc not in mapping[qe_group]:
            mapping[qe_group][pc] = []
        mapping[qe_group][pc].append({
            # NOTE: must normalize here, at first ingestion, while the
            # *original* string length from the spreadsheet is still known.
            # A blind .zfill(8) here would already destroy the information
            # _normalize_itchs() needs (see its docstring) before any
            # downstream code gets a chance to look at it.
            "itchs": _normalize_itchs(itchs),
            "unit":  unit,
            "desc":  desc,
        })
    return mapping


def _normalize_itchs(raw):
    """
    The QE/PC mapping file stores ITCHS codes as 8 digits. Excel cells that
    happen to be typed as a *number* instead of text silently drop leading
    zeros when read back (e.g. true code "08401000" -> "8401000"). In the
    mapping file this only ever costs a single digit in practice — the
    7-digit case below — because every HS chapter that can lose a leading
    zero this way is 01-09, so restoring exactly one zero is unambiguous.
    A small number of source rows are genuinely shorter/malformed (e.g. one
    stray 6-digit entry in the mapping file). For those we don't know how
    many digits, if any, are missing at the front, so we trust the leading
    digits as typed — that's what determines the HS chapter, and inventing
    extra leading zeros would fabricate a chapter (e.g. "00") that doesn't
    exist in the HS nomenclature. Either way, return a string at least
    `len` long by padding on the RIGHT (never the left) for finer
    granularity — right-padding only fills in *trailing* digits we
    genuinely don't have, it never changes the chapter/heading.
    """
    raw = raw.strip()
    if len(raw) == 7 and raw.isdigit():
        raw = raw.zfill(8)
    return raw


def extract_hs_codes(pc_data, hs_level=2):
    """
    From a dict of {pc_group: [{"itchs": ..., "desc": ...}]},
    extract unique HS codes at the requested level (2, 4, 6, or 8 digits).
    Returns: {hs_code: {"pc_groups": set(), "descriptions": set()}}
    """
    result = {}
    for pc, codes in pc_data.items():
        for entry in codes:
            raw = _normalize_itchs(entry["itchs"])
            hs  = raw[:hs_level].ljust(hs_level, "0") if hs_level <= 8 else raw

            if hs not in result:
                result[hs] = {"pc_groups": set(), "descriptions": set()}
            result[hs]["pc_groups"].add(pc)
            # Use first 50 chars of description
            if entry["desc"]:
                result[hs]["descriptions"].add(entry["desc"][:60])
    return result


def bucket_for_code(hs_code, pc_data, hs_level=2):
    """Return the primary PC group (bucket) for an HS code."""
    for pc, codes in pc_data.items():
        for entry in codes:
            raw = _normalize_itchs(entry["itchs"])
            candidate = raw[:hs_level].ljust(hs_level, "0")
            if candidate == hs_code:
                return pc
    return "Other"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LIST MODE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def list_groups(mapping):
    print("\n  Available QE Groups (Major Commodity Groups)")
    print("  " + "─" * 55)
    for i, group in enumerate(sorted(mapping.keys()), 1):
        pc_count   = len(mapping[group])
        code_count = sum(len(v) for v in mapping[group].values())
        print(f"  {i:>2}.  {group:<50}  ({pc_count} PC groups, {code_count} ITCHS codes)")
    print()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# KEYWORD SEARCH
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def search_groups(mapping, keyword):
    kw = keyword.lower()
    matches = [g for g in mapping if kw in g.lower()]
    if not matches:
        print(f"\n  No QE groups found matching '{keyword}'.")
        print("  Use --list to see all available groups.\n")
        sys.exit(0)
    print(f"\n  Groups matching '{keyword}':")
    print("  " + "─" * 50)
    for m in matches:
        pc_count   = len(mapping[m])
        code_count = sum(len(v) for v in mapping[m].values())
        print(f"  → {m}  ({pc_count} PC groups, {code_count} ITCHS codes)")
    print()
    if len(matches) == 1:
        return matches[0]
    return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIG BUILDER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def slugify(text):
    import re
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def merge_pc_data(mapping, group_names):
    """
    Merge pc_data from multiple QE groups into one combined dict.
    If the same PC group name appears in multiple QE groups, they are merged.
    Returns {pc_group: [entries]}, source_labels list.
    """
    combined = {}
    for gname in group_names:
        match = next((g for g in mapping if g.upper() == gname.strip().upper()), None)
        if not match:
            print(f"  ⚠  Group not found (skipped): '{gname.strip()}'")
            continue
        for pc, entries in mapping[match].items():
            # Prefix PC group with QE group name to keep buckets distinct
            label = f"{match.title()} — {pc}"
            if label not in combined:
                combined[label] = []
            combined[label].extend(entries)
    return combined


def build_config(
    sector_name,
    pc_data=None,        # {pc_group: [{"itchs":..., "desc":...}]} or None for manual
    hs_level=2,
    hs_filter=None,      # set of 2-digit chapters to include; None = include all
    extra_codes=None,    # list of extra HS codes to add
    out_path=None,
    flow="export",
    source_group=None,   # original QE group name(s) for documentation
):
    """
    Build and save a sector_config.xlsx.
    If pc_data is None, generates a blank config for manual entry.
    """
    TODAY    = str(date.today())
    slug     = slugify(sector_name)
    out_path = out_path or Path(f"{slug}_sector_config.xlsx")

    # ── Gather HS codes from QE/PC mapping ───────────────────
    hs_code_rows = []   # list of (hs_code, description, bucket, source)

    if pc_data:
        hs_info = extract_hs_codes(pc_data, hs_level)
        for hs_code in sorted(hs_info.keys()):
            # Apply hs_filter: skip codes whose 2-digit chapter is not in the filter
            if hs_filter and hs_code[:2].zfill(2) not in hs_filter:
                continue
            info    = hs_info[hs_code]
            buckets = sorted(info["pc_groups"])
            bucket  = buckets[0] if buckets else "Other"
            desc    = "; ".join(sorted(info["descriptions"]))[:80] if info["descriptions"] else ""
            hs_code_rows.append((hs_code, desc, bucket, "QE/PC Auto"))

    # ── Add extra manual codes ────────────────────────────────
    existing_codes = {r[0] for r in hs_code_rows}
    if extra_codes:
        for code in extra_codes:
            code = code.strip().zfill(hs_level if hs_level <= 8 else 2)
            if code not in existing_codes:
                hs_code_rows.append((code, "", "Custom", "Manual Addition"))
                existing_codes.add(code)

    # ── Derive buckets from PC groups ─────────────────────────
    bucket_rows = []
    if pc_data:
        for pc_group in sorted(pc_data.keys()):
            codes_in_bucket = sorted(set(
                r[0] for r in hs_code_rows if r[2] == pc_group
            ))
            bucket_rows.append((pc_group, codes_in_bucket))
    if any(r[2] == "Custom" for r in hs_code_rows):
        custom_codes = sorted(r[0] for r in hs_code_rows if r[2] == "Custom")
        bucket_rows.append(("Custom Group", custom_codes))

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # BUILD WORKBOOK
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    wb = openpyxl.Workbook()

    # ── Sheet 1: sector_details ───────────────────────────────
    ws1 = wb.active
    ws1.title = "sector_details"
    write_title(ws1, 1, "sector_details", 2)
    write_header(ws1, 2, [("Field", 35), ("Value", 55)])
    details = [
        ("Sector name",            sector_name),
        ("Sector slug",            slug),
        ("Sector description",     ""),
        ("Country focus",          "India"),
        ("Source database",        "UN Comtrade (comtradeapicall)"),
        ("Currency unit",          "USD Mn"),
        ("QE source group",        source_group or "Manual"),
        ("Flow used for mapping",  flow.upper()),
        ("HS code level",          f"{hs_level}-digit"),
        ("Created date",           TODAY),
    ]
    for i, (f, v) in enumerate(details, 3):
        style_cell(ws1, i, 1, f, bold=True, fill=LIGHT)
        style_cell(ws1, i, 2, v, wrap=True)
    note_row(ws1, len(details) + 3, "← Fill in Sector description before running the pipeline.", 2)

    # ── Sheet 2: hs_codes ─────────────────────────────────────
    ws2 = wb.create_sheet("hs_codes")
    title_text = f"hs_codes — {sector_name} ({'auto-filled from QE/PC' if pc_data else 'manual entry'})"
    write_title(ws2, 1, title_text, 6)
    write_header(ws2, 2, [
        ("HS Code",               12),
        ("Product Description",   50),
        ("Include / Exclude",     18),
        ("Sector Bucket",         28),
        ("Source",                18),
        ("Notes",                 25),
    ])
    note_row(ws2, 3,
             "← HS codes are text. Only rows marked 'Include' are downloaded. "
             "Change 'Include' to 'Exclude' to skip a code. Add rows manually at the bottom.",
             6)

    if hs_code_rows:
        for i, (code, desc, bucket, source) in enumerate(hs_code_rows, 4):
            fill = LIGHT if i % 2 == 0 else None
            src_fill = GOLD if source == "Manual Addition" else fill
            style_cell(ws2, i, 1, code,      fill=fill)
            style_cell(ws2, i, 2, desc,      fill=fill, wrap=True)
            style_cell(ws2, i, 3, "Include", fill=fill)
            style_cell(ws2, i, 4, bucket,    fill=fill)
            style_cell(ws2, i, 5, source,    fill=src_fill)
            style_cell(ws2, i, 6, "",        fill=fill)
            ws2.row_dimensions[i].height = 20
    else:
        # Blank rows for manual entry
        note_row(ws2, 4, "← Add your HS codes below. One row per code.", 6)
        for i in range(5, 15):
            for c in range(1, 7):
                style_cell(ws2, i, c, "")

    # ── Sheet 3: sector_buckets ───────────────────────────────
    ws3 = wb.create_sheet("sector_buckets")
    write_title(ws3, 1, f"sector_buckets — {sector_name}", 4)
    write_header(ws3, 2, [
        ("Sector Bucket",   30),
        ("Definition",      50),
        ("HS Codes",        28),
        ("Notes",           25),
    ])
    if bucket_rows:
        for i, (bucket, codes) in enumerate(bucket_rows, 3):
            fill = LIGHT if i % 2 == 1 else None
            style_cell(ws3, i, 1, bucket,           bold=True, fill=fill)
            style_cell(ws3, i, 2, "",               fill=fill, wrap=True)
            style_cell(ws3, i, 3, ",".join(codes),  fill=fill)
            style_cell(ws3, i, 4, "",               fill=fill)
            ws3.row_dimensions[i].height = 22
        note_row(ws3, len(bucket_rows) + 3,
                 "← Fill in the Definition column before running the pipeline.", 4)
    else:
        note_row(ws3, 3, "← Add sector buckets (groupings) here. One bucket per row.", 4)
        for i in range(4, 12):
            for c in range(1, 5):
                style_cell(ws3, i, c, "")

    # ── Sheet 4: years ────────────────────────────────────────
    ws4 = wb.create_sheet("years")
    write_title(ws4, 1, "years — Time window for data download", 3)
    write_header(ws4, 2, [("Start Year", 14), ("End Year", 14), ("Period String", 60)])
    style_cell(ws4, 3, 1, 2014, fill=LIGHT)
    style_cell(ws4, 3, 2, 2023, fill=LIGHT)
    style_cell(ws4, 3, 3, "2014,2015,2016,2017,2018,2019,2020,2021,2022,2023", fill=LIGHT)
    note_row(ws4, 4, "← Change Start Year / End Year and update Period String (comma-separated).", 3)

    # ── Sheet 5: flows ────────────────────────────────────────
    ws5 = wb.create_sheet("flows")
    write_title(ws5, 1, "flows — 4 standard Comtrade trade flows", 7)
    write_header(ws5, 2, [
        ("Flow Name",     20), ("Flow Code", 12), ("Reporter", 15),
        ("Reporter Code", 15), ("Partner",   15), ("Partner Code", 14),
        ("Output File",   32),
    ])
    default_flows = [
        ("World Exports", "X", "All (None)", "None", "World", "0", "out_world_exports.xlsx"),
        ("World Imports", "M", "All (None)", "None", "World", "0", "out_world_imports.xlsx"),
        ("India Exports", "X", "India",      "699",  "World", "0", "out_india_exports.xlsx"),
        ("India Imports", "M", "India",      "699",  "World", "0", "out_india_imports.xlsx"),
    ]
    for i, row in enumerate(default_flows, 3):
        fill = LIGHT if i % 2 == 1 else None
        for c, v in enumerate(row, 1):
            style_cell(ws5, i, c, v, fill=fill)
    note_row(ws5, 7, "← Usually no changes needed. To change the country focus, update Reporter / Reporter Code.", 7)

    # ── Sheet 6: settings ─────────────────────────────────────
    ws6 = wb.create_sheet("settings")
    write_title(ws6, 1, "settings — Pipeline configuration", 2)
    write_header(ws6, 2, [("Setting", 42), ("Value", 38)])
    test_hs = hs_code_rows[0][0] if hs_code_rows else "[Enter one HS code for test]"
    settings = [
        ("API key environment variable name", "COMTRADE_API_KEY"),
        ("Test HS code",                      test_hs),
        ("Test year",                         "2023"),
        ("Output folder",                     "."),
        ("Error log filename",                f"{slug}_errors.log"),
        ("Continue on error",                 "Yes"),
        ("Overwrite existing output files",   "Yes"),
    ]
    for i, (s, v) in enumerate(settings, 3):
        style_cell(ws6, i, 1, s, bold=True, fill=LIGHT)
        style_cell(ws6, i, 2, v)
    note_row(ws6, len(settings) + 3,
             "← Set COMTRADE_API_KEY as an env variable before running. "
             "See pipeline script for instructions.", 2)

    # ── Sheet 7: custom_groups ────────────────────────────────
    ws7 = wb.create_sheet("custom_groups")
    write_title(ws7, 1,
                "custom_groups — Add your own HS codes here (manual override)",
                fill=MID, ncols=6)
    write_header(ws7, 2, [
        ("HS Code",           12),
        ("Product Description", 50),
        ("Sector Bucket",     28),
        ("Reason / Note",     40),
        ("Include / Exclude", 18),
        ("Source",            20),
    ], fill=MID)
    note_row(ws7, 3,
             "← Add any HS codes that are NOT in the QE/PC mapping, or that you want to add "
             "to a specific bucket manually. Then run:  python create_sector_config_from_qe.py "
             f"--group \"{source_group or sector_name}\" --merge_custom {out_path.name}", 6)
    note_row(ws7, 4,
             "   OR simply copy these rows into the hs_codes sheet directly before running "
             "the pipeline.", 6)
    for i in range(5, 20):
        fill = LIGHT if i % 2 == 1 else None
        for c in range(1, 7):
            style_cell(ws7, i, c, "", fill=fill)
        style_cell(ws7, i, 5, "Include", fill=fill)
        style_cell(ws7, i, 6, "Manual", fill=fill)
        ws7.row_dimensions[i].height = 20

    # ── Sheet 8: source_log ───────────────────────────────────
    ws8 = wb.create_sheet("source_log")
    write_title(ws8, 1, "source_log — How this config was generated", 2, fill="555555")
    write_header(ws8, 2, [("Item", 35), ("Value", 60)], fill="555555")
    log_items = [
        ("Generated by",           "create_sector_config_from_qe.py"),
        ("Generation date",        TODAY),
        ("Sector name",            sector_name),
        ("QE source group",        source_group or "Manual"),
        ("Mapping file used",      "QE_PC_HS_Mapping_2025-26_Final_Sent.xlsx"),
        ("Flow sheet used",        flow.upper()),
        ("HS code level",          f"{hs_level}-digit"),
        ("Total HS codes loaded",  str(len(hs_code_rows))),
        ("Total buckets",          str(len(bucket_rows))),
        ("Extra codes added",      ",".join(extra_codes) if extra_codes else "None"),
    ]
    for i, (k, v) in enumerate(log_items, 3):
        style_cell(ws8, i, 1, k, bold=True, fill=GRAY)
        style_cell(ws8, i, 2, v)

    wb.save(out_path)
    return out_path


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MERGE CUSTOM GROUPS INTO HS_CODES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def merge_custom_into_hs_codes(config_path):
    """
    Reads custom_groups sheet and appends any new codes to hs_codes sheet.
    Skips codes that are already present in hs_codes.
    """
    wb = openpyxl.load_workbook(config_path)
    if "custom_groups" not in wb.sheetnames:
        print("  ⚠  No custom_groups sheet found.")
        return
    if "hs_codes" not in wb.sheetnames:
        print("  ⚠  No hs_codes sheet found.")
        return

    ws_custom = wb["custom_groups"]
    ws_hs     = wb["hs_codes"]

    # Find existing codes in hs_codes
    existing = set()
    for row in ws_hs.iter_rows(min_row=4, values_only=True):
        if row[0]:
            existing.add(str(row[0]).strip())

    # Find last row in hs_codes
    last_row = ws_hs.max_row + 1

    added = 0
    for row in ws_custom.iter_rows(min_row=5, values_only=True):
        hs_code     = str(row[0]).strip() if row[0] else None
        desc        = str(row[1]).strip() if row[1] else ""
        bucket      = str(row[2]).strip() if row[2] else "Custom"
        note        = str(row[3]).strip() if row[3] else ""
        include     = str(row[4]).strip() if row[4] else "Include"

        if not hs_code or hs_code in existing:
            continue

        i = last_row + added
        fill = LIGHT if i % 2 == 0 else None
        style_cell(ws_hs, i, 1, hs_code,  fill=fill)
        style_cell(ws_hs, i, 2, desc,     fill=fill, wrap=True)
        style_cell(ws_hs, i, 3, include,  fill=fill)
        style_cell(ws_hs, i, 4, bucket,   fill=fill)
        style_cell(ws_hs, i, 5, "Manual Addition", fill=fill)
        style_cell(ws_hs, i, 6, note,     fill=fill)
        existing.add(hs_code)
        added += 1

    if added:
        wb.save(config_path)
        print(f"  ✅  Merged {added} custom code(s) into hs_codes sheet → {config_path}")
    else:
        print("  ✓  No new codes to merge (all already present in hs_codes).")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    parser = argparse.ArgumentParser(
        description="Generate sector config from QE/PC HS mapping",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--group",        type=str, help="Single exact QE group name")
    parser.add_argument("--groups",       type=str, help="Comma-separated QE group names to combine")
    parser.add_argument("--keyword",      type=str, help="Search keyword for QE group")
    parser.add_argument("--manual",       action="store_true", help="Generate blank config for manual entry")
    parser.add_argument("--sector",       type=str, help="Sector name (required with --groups or --manual)")
    parser.add_argument("--flow",         type=str, default="export", choices=["export","import"],
                        help="Use EXPORT or IMPORT sheet (default: export)")
    parser.add_argument("--hs_level",     type=int, default=2, choices=[2, 4, 6, 8],
                        help="HS code granularity: 2=chapter, 6=sub-heading, 8=ITCHS (default: 2)")
    parser.add_argument("--hs_filter",    type=str,
                        help="Comma-separated 2-digit HS chapters to include (e.g. '50,51,52'). All others excluded.")
    parser.add_argument("--add_codes",    type=str, help="Comma-separated extra HS codes to add")
    parser.add_argument("--out",          type=str, help="Output file name")
    parser.add_argument("--merge_custom", type=str, metavar="CONFIG_FILE",
                        help="Merge custom_groups into hs_codes of an existing config file")
    parser.add_argument("--list",         action="store_true", help="List all available QE groups")
    parser.add_argument("--mapping",      type=str, help="Path to QE/PC mapping file")

    args = parser.parse_args()

    # ── Resolve mapping file ──────────────────────────────────
    mapping_path = Path(args.mapping) if args.mapping else DEFAULT_MAPPING
    if not mapping_path.exists():
        print(f"\n  ✗  Mapping file not found: {mapping_path}")
        print("     Place QE_PC_HS_Mapping_2025-26_Final_Sent.xlsx in the same folder,")
        print("     or specify its location with --mapping path/to/file.xlsx\n")
        sys.exit(1)

    # ── Merge mode ────────────────────────────────────────────
    if args.merge_custom:
        merge_custom_into_hs_codes(Path(args.merge_custom))
        return

    # ── Load mapping ──────────────────────────────────────────
    print(f"\n  Loading mapping from {mapping_path.name} ...")
    mapping = load_mapping(mapping_path, flow=args.flow)
    print(f"  ✓  Loaded {len(mapping)} QE groups.")

    # ── List mode ─────────────────────────────────────────────
    if args.list:
        list_groups(mapping)
        return

    # ── Manual mode ───────────────────────────────────────────
    if args.manual:
        if not args.sector:
            print("\n  ✗  --sector is required with --manual.")
            print("     Example: python create_sector_config_from_qe.py --manual --sector 'Toys and Games'\n")
            sys.exit(1)
        extra = [c.strip() for c in args.add_codes.split(",")] if args.add_codes else []
        out_path = Path(args.out) if args.out else Path(f"{slugify(args.sector)}_sector_config.xlsx")
        saved = build_config(
            sector_name  = args.sector,
            pc_data      = None,
            hs_level     = args.hs_level,
            extra_codes  = extra,
            out_path     = out_path,
            flow         = args.flow,
            source_group = None,
        )
        print(f"\n  ✅  Blank config saved → {saved}")
        print(f"      Open it in Excel and fill in the hs_codes and sector_buckets sheets.")
        print(f"      Then run the pipeline:\n")
        print(f"      python sector_comtrade_pipeline.py --config {saved.name} --mode setup_check\n")
        return

    # ── Parse hs_filter ───────────────────────────────────────
    hs_filter = None
    if args.hs_filter:
        hs_filter = {c.strip().zfill(2) for c in args.hs_filter.split(",")}
        print(f"  HS filter active: only chapters {sorted(hs_filter)} will be included.")

    # ── Multi-group mode ──────────────────────────────────────
    if args.groups:
        if not args.sector:
            print("\n  ✗  --sector is required with --groups.")
            print("     Example: --groups 'RMG OF ALL TEXTILES,CARPET' --sector 'Textiles & Apparel'\n")
            sys.exit(1)
        group_names = [g.strip() for g in args.groups.split(";") if g.strip()]
        print(f"\n  Combining {len(group_names)} QE groups for sector: {args.sector}")
        for gn in group_names:
            print(f"    • {gn}")

        combined_pc = merge_pc_data(mapping, group_names)
        if not combined_pc:
            # All requested groups failed to match the mapping file. Writing
            # a config here would silently produce a 0-code sector that
            # looks "successful" downstream (clean/visuals would run but
            # output nothing) — fail loudly instead so the bad group name(s)
            # get fixed at the source.
            print(f"\n  ✗  None of the {len(group_names)} requested group(s) matched the "
                  f"mapping file — no config was written.")
            print("     Group matching is exact (case-insensitive). Use --list to see the "
                  "exact wording of every Major Commodity Group.\n")
            sys.exit(1)
        extra       = [c.strip() for c in args.add_codes.split(",")] if args.add_codes else []
        out_path    = Path(args.out) if args.out else Path(f"{slugify(args.sector)}_sector_config.xlsx")

        # Preview counts
        hs_info     = extract_hs_codes(combined_pc, args.hs_level)
        if hs_filter:
            hs_info = {k: v for k, v in hs_info.items() if k[:2].zfill(2) in hs_filter}
        print(f"\n  Combined PC groups : {len(combined_pc)}")
        print(f"  Unique {args.hs_level}-digit HS codes (after filter): {len(hs_info)}")
        print(f"  Output: {out_path}\n")

        saved = build_config(
            sector_name  = args.sector,
            pc_data      = combined_pc,
            hs_level     = args.hs_level,
            hs_filter    = hs_filter,
            extra_codes  = extra,
            out_path     = out_path,
            flow         = args.flow,
            source_group = "; ".join(group_names),
        )
        print(f"  ✅  Config saved → {saved}")
        print(f"\n  NEXT STEPS:")
        print(f"  1. Open {saved.name} in Excel")
        print(f"  2. Fill in Sector description (sector_details sheet)")
        print(f"  3. Fill in bucket Definitions (sector_buckets sheet)")
        print(f"  4. Review hs_codes — change 'Include' to 'Exclude' for any codes you don't want")
        print(f"  5. Run: python sector_comtrade_pipeline.py --config {saved.name} --mode setup_check\n")
        return

    # ── Group / keyword mode ──────────────────────────────────
    group = args.group
    if args.keyword and not group:
        group = search_groups(mapping, args.keyword)
        if not group:
            print("  Multiple matches found. Re-run with --group using the exact name above.\n")
            return

    if not group:
        parser.print_help()
        print("\n  ✗  Specify --group, --keyword, --manual, or --list.\n")
        sys.exit(1)

    # Exact match (case-insensitive)
    group_match = next((g for g in mapping if g.upper() == group.upper()), None)
    if not group_match:
        print(f"\n  ✗  Group '{group}' not found.")
        print("     Use --list to see all groups, or --keyword for fuzzy search.\n")
        sys.exit(1)

    pc_data  = mapping[group_match]
    slug     = slugify(group_match)
    extra    = [c.strip() for c in args.add_codes.split(",")] if args.add_codes else []
    out_path = Path(args.out) if args.out else Path(f"{slug}_sector_config.xlsx")

    # Summary before building
    total_itchs = sum(len(v) for v in pc_data.values())
    hs_info     = extract_hs_codes(pc_data, args.hs_level)
    print(f"\n  QE Group  : {group_match}")
    print(f"  PC Groups : {len(pc_data)}")
    print(f"  ITCHS codes in mapping : {total_itchs}")
    print(f"  Unique {args.hs_level}-digit HS codes : {len(hs_info)}")
    if extra:
        print(f"  Extra codes to add : {extra}")
    print(f"  Output file : {out_path}\n")

    saved = build_config(
        sector_name  = group_match.title(),
        pc_data      = pc_data,
        hs_level     = args.hs_level,
        hs_filter    = hs_filter,
        extra_codes  = extra,
        out_path     = out_path,
        flow         = args.flow,
        source_group = group_match,
    )
    print(f"  ✅  Config saved → {saved}")
    print(f"\n  NEXT STEPS:")
    print(f"  1. Open {saved.name} in Excel")
    print(f"  2. Fill in Sector description (sector_details sheet)")
    print(f"  3. Fill in bucket Definitions (sector_buckets sheet)")
    print(f"  4. Review hs_codes — change 'Include' to 'Exclude' for any codes you don't want")
    print(f"  5. Add any extra codes to the custom_groups sheet if needed")
    print(f"  6. To merge custom_groups into hs_codes, run:")
    print(f"       python create_sector_config_from_qe.py --merge_custom {saved.name}")
    print(f"  7. Run the pipeline:")
    print(f"       python sector_comtrade_pipeline.py --config {saved.name} --mode setup_check\n")


if __name__ == "__main__":
    main()
